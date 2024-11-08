from openpyxl import load_workbook
from tqdm import tqdm


def extract_formulas(filename):
    # Load the workbook
    wb = load_workbook(filename)

    # Create a new worksheet for storing the formulas and insert it at the beginning of the workbook
    new_ws = wb.create_sheet(title="Formulas Index", index=0)

    # Add header row to the new worksheet
    new_ws.append(["Worksheet", "Cell", "Formula"])

    num_formulas_total = 0

    # Loop through each worksheet in the workbook
    for ws in wb.worksheets:
        num_formulas = 0

        # Count the total number of rows in the worksheet
        total_rows = ws.max_row
        total_columns = ws.max_column

        # Initialize tqdm progress bar
        progress_bar = tqdm(
            desc=f"Processing worksheet '{ws.title}'", total=ws.max_row, ncols=120
        )

        for row in range(1, total_rows + 1):
            for col in range(1, total_columns + 1):
                cell = ws.cell(row=row, column=col)
                if cell.data_type == "f":
                    num_formulas += 1
                    # Write worksheet name, cell address, and formula into the new worksheet
                    try:
                        new_ws.append([ws.title, cell.coordinate, cell.value])
                    except ValueError as e:
                        new_ws.append([ws.title, cell.coordinate, "ERROR"])

            progress_bar.update(1)

        progress_bar.close()

        num_formulas_total += num_formulas
        print(f"{ws.title} contained {num_formulas:,} formulas.")

    print(f"{num_formulas_total:,} formulas found in workbook.")

    # Save the workbook
    wb.save(filename)


extract_formulas(r"C:\Users\Documents\YourExcelFile.xlsx")
